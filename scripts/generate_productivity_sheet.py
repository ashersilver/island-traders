"""
Generate island_traders/board/productivity.xlsx

One tab per island showing seasonal inputs → outputs, with editable yellow cells
for quantities. Run this whenever you want a fresh copy of the sheet:

    python scripts/generate_productivity_sheet.py
"""
from pathlib import Path
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

OUT = Path(__file__).parent.parent / "island_traders" / "board" / "productivity.xlsx"

SEASONS = ["Spring", "Summer", "Autumn", "Winter"]
RESOURCES = [
    "Food", "Fish", "Ore", "Oil", "Freight",
    "Knowledge", "CapitalEquipment", "Goods",
    "HealthServices", "Vaccine", "Finance",
]

# ── Colour palette ──────────────────────────────────────────────────────────
C_HEADER_BG   = "1C1410"   # dark ink
C_HEADER_FG   = "F5EFE6"   # cream
C_SECTION_BG  = "3D2B1A"   # mid-dark
C_SEASON_BG   = {
    "Spring": "D4EDBA", "Summer": "FFF3CD",
    "Autumn": "FFE0B2", "Winter": "BBDEFB",
}
C_INPUT_BG    = "FFF9E6"   # pale yellow — editable cells
C_OUTPUT_BG   = "E8F5E9"   # pale green — editable cells
C_YIELD_BG    = "E3F2FD"   # pale blue — seasonal yield
C_LABOUR_BG   = "FCE4EC"   # pale pink — labour
C_CALC_BG     = "F5F5F5"   # grey — formula cells (computed)
C_BORDER      = "9E9B94"

THIN = Side(style="thin", color=C_BORDER)
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def hdr_font(bold=True):
    return Font(name="Calibri", bold=bold, color=C_HEADER_FG, size=10)

def body_font(bold=False, color="1C1410"):
    return Font(name="Calibri", bold=bold, color=color, size=10)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def centre():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def write_cell(ws, row, col, value, bg=None, bold=False, fg="1C1410",
               align=None, num_fmt=None, border=True):
    c = ws.cell(row=row, column=col, value=value)
    if bg:
        c.fill = fill(bg)
    c.font = Font(name="Calibri", bold=bold, color=fg, size=10)
    c.alignment = align or centre()
    if num_fmt:
        c.number_format = num_fmt
    if border:
        c.border = BORDER
    return c


# ── Island definitions ──────────────────────────────────────────────────────
# Each island entry: name, role_key, seasonal_yield, labour requirements,
# and per-season {inputs, outputs}.

ISLANDS = [
    {
        "name": "Greenfeld — Agriculture & Fisheries",
        "role": "Farmer",
        "desc": "Table-based seasonal conversion. Edit yellow cells to tune inputs/outputs.",
        "seasonal_yield": None,   # governed by table, not a multiplier
        "labour": {
            "Spring":  {"Skilled": 1, "Unskilled": 3},
            "Summer":  {"Skilled": 1, "Unskilled": 5},
            "Autumn":  {"Skilled": 2, "Unskilled": 6},
            "Winter":  {"Skilled": 1, "Unskilled": 1},
        },
        "seasons": {
            "Spring": {
                "inputs":  {"CapitalEquipment": 1, "Oil": 1},
                "outputs": {"Food": 2, "Fish": 3},
            },
            "Summer": {
                "inputs":  {"CapitalEquipment": 1, "Oil": 1},
                "outputs": {"Food": 3, "Fish": 5},
            },
            "Autumn": {
                "inputs":  {"CapitalEquipment": 1, "Oil": 1},
                "outputs": {"Food": 7, "Fish": 2},
            },
            "Winter": {
                "inputs":  {"CapitalEquipment": 1, "Oil": 1},
                "outputs": {"Food": 2, "Fish": 1},
            },
        },
    },
    {
        "name": "Ironvein Isle — Mining & Oil",
        "role": "Miner",
        "desc": "Seasonal yield multiplier applied to base outputs. Edit yellow cells.",
        "seasonal_yield": {"Spring": 1.0, "Summer": 1.1, "Autumn": 1.0, "Winter": 0.9},
        "labour": {
            "Spring":  {"Skilled": 2, "Unskilled": 3},
            "Summer":  {"Skilled": 2, "Unskilled": 3},
            "Autumn":  {"Skilled": 2, "Unskilled": 3},
            "Winter":  {"Skilled": 2, "Unskilled": 2},
        },
        "seasons": {
            "Spring": {"inputs": {"Oil": 1, "Freight": 1}, "outputs": {"Ore": 5, "Oil": 3}},
            "Summer": {"inputs": {"Oil": 1, "Freight": 1}, "outputs": {"Ore": 5, "Oil": 3}},
            "Autumn": {"inputs": {"Oil": 1, "Freight": 1}, "outputs": {"Ore": 5, "Oil": 3}},
            "Winter": {"inputs": {"Oil": 1, "Freight": 1}, "outputs": {"Ore": 5, "Oil": 3}},
        },
    },
    {
        "name": "Port Sovereign — Transportation & Shipping",
        "role": "Transporter",
        "desc": "Seasonal yield multiplier. Winter storms reduce output. Edit yellow cells.",
        "seasonal_yield": {"Spring": 0.9, "Summer": 1.1, "Autumn": 1.2, "Winter": 0.8},
        "labour": {
            "Spring":  {"Skilled": 2, "Unskilled": 2},
            "Summer":  {"Skilled": 2, "Unskilled": 3},
            "Autumn":  {"Skilled": 2, "Unskilled": 4},
            "Winter":  {"Skilled": 1, "Unskilled": 2},
        },
        "seasons": {
            "Spring": {"inputs": {"Oil": 2, "CapitalEquipment": 1}, "outputs": {"Freight": 6}},
            "Summer": {"inputs": {"Oil": 2, "CapitalEquipment": 1}, "outputs": {"Freight": 6}},
            "Autumn": {"inputs": {"Oil": 2, "CapitalEquipment": 1}, "outputs": {"Freight": 6}},
            "Winter": {"inputs": {"Oil": 2, "CapitalEquipment": 1}, "outputs": {"Freight": 6}},
        },
    },
    {
        "name": "Scholara — Education & Training",
        "role": "Educator",
        "desc": "Summer break reduces output. Edit yellow cells.",
        "seasonal_yield": {"Spring": 1.1, "Summer": 0.7, "Autumn": 1.1, "Winter": 1.1},
        "labour": {
            "Spring":  {"Skilled": 2, "Unskilled": 2},
            "Summer":  {"Skilled": 1, "Unskilled": 1},
            "Autumn":  {"Skilled": 2, "Unskilled": 2},
            "Winter":  {"Skilled": 2, "Unskilled": 2},
        },
        "seasons": {
            "Spring": {"inputs": {"CapitalEquipment": 1, "Finance": 1}, "outputs": {"Knowledge": 4}},
            "Summer": {"inputs": {"CapitalEquipment": 1, "Finance": 1}, "outputs": {"Knowledge": 4}},
            "Autumn": {"inputs": {"CapitalEquipment": 1, "Finance": 1}, "outputs": {"Knowledge": 4}},
            "Winter": {"inputs": {"CapitalEquipment": 1, "Finance": 1}, "outputs": {"Knowledge": 4}},
        },
    },
    {
        "name": "Aurum Cay — Banking",
        "role": "Banker",
        "desc": "Economic cycle aligns with harvest/trade seasons. Edit yellow cells.",
        "seasonal_yield": {"Spring": 0.9, "Summer": 1.0, "Autumn": 1.2, "Winter": 0.9},
        "labour": {
            "Spring":  {"Skilled": 2, "Unskilled": 1},
            "Summer":  {"Skilled": 2, "Unskilled": 1},
            "Autumn":  {"Skilled": 2, "Unskilled": 2},
            "Winter":  {"Skilled": 2, "Unskilled": 1},
        },
        "seasons": {
            "Spring": {"inputs": {"Knowledge": 1, "CapitalEquipment": 1}, "outputs": {"Finance": 3}},
            "Summer": {"inputs": {"Knowledge": 1, "CapitalEquipment": 1}, "outputs": {"Finance": 3}},
            "Autumn": {"inputs": {"Knowledge": 1, "CapitalEquipment": 1}, "outputs": {"Finance": 3}},
            "Winter": {"inputs": {"Knowledge": 1, "CapitalEquipment": 1}, "outputs": {"Finance": 3}},
        },
    },
    {
        "name": "Forgehaven — Manufacturing",
        "role": "Manufacturer",
        "desc": "Demand-driven production. Autumn peak follows harvest. Edit yellow cells.",
        "seasonal_yield": {"Spring": 1.0, "Summer": 1.0, "Autumn": 1.1, "Winter": 0.9},
        "labour": {
            "Spring":  {"Skilled": 3, "Unskilled": 2},
            "Summer":  {"Skilled": 3, "Unskilled": 2},
            "Autumn":  {"Skilled": 3, "Unskilled": 2},
            "Winter":  {"Skilled": 2, "Unskilled": 2},
        },
        "seasons": {
            "Spring": {"inputs": {"Ore": 2, "Oil": 1, "Freight": 1}, "outputs": {"Goods": 3, "CapitalEquipment": 2}},
            "Summer": {"inputs": {"Ore": 2, "Oil": 1, "Freight": 1}, "outputs": {"Goods": 3, "CapitalEquipment": 2}},
            "Autumn": {"inputs": {"Ore": 2, "Oil": 1, "Freight": 1}, "outputs": {"Goods": 3, "CapitalEquipment": 2}},
            "Winter": {"inputs": {"Ore": 2, "Oil": 1, "Freight": 1}, "outputs": {"Goods": 3, "CapitalEquipment": 2}},
        },
    },
    {
        "name": "Mericula — Healthcare",
        "role": "Doctor",
        "desc": "Winter illness and Summer injuries drive demand peaks. Edit yellow cells.",
        "seasonal_yield": {"Spring": 0.9, "Summer": 1.1, "Autumn": 0.9, "Winter": 1.1},
        "labour": {
            "Spring":  {"Skilled": 12, "Unskilled": 18},
            "Summer":  {"Skilled": 14, "Unskilled": 21},
            "Autumn":  {"Skilled": 12, "Unskilled": 18},
            "Winter":  {"Skilled": 24, "Unskilled": 20},
        },
        "seasons": {
            "Spring": {"inputs": {"Knowledge": 1, "CapitalEquipment": 1}, "outputs": {"HealthServices": 4, "Vaccine": 1}},
            "Summer": {"inputs": {"Knowledge": 1, "CapitalEquipment": 1}, "outputs": {"HealthServices": 4, "Vaccine": 1}},
            "Autumn": {"inputs": {"Knowledge": 1, "CapitalEquipment": 1}, "outputs": {"HealthServices": 4, "Vaccine": 1}},
            "Winter": {"inputs": {"Knowledge": 1, "CapitalEquipment": 1}, "outputs": {"HealthServices": 4, "Vaccine": 1}},
        },
    },
]

BASE_PRICES = {
    "Food": 10, "Fish": 8, "Ore": 15, "Oil": 20, "Freight": 12,
    "Knowledge": 18, "CapitalEquipment": 28, "Goods": 30,
    "HealthServices": 35, "Vaccine": 40, "Finance": 20,
}


def build_island_sheet(wb, island: dict):
    ws = wb.create_sheet(title=island["role"])
    ws.freeze_panes = "B4"

    # Column widths
    ws.column_dimensions["A"].width = 26
    for col in range(2, 14):
        ws.column_dimensions[get_column_letter(col)].width = 14

    # ── Row 1: Island name ─────────────────────────────────────
    ws.merge_cells("A1:M1")
    c = ws["A1"]
    c.value = island["name"]
    c.font = Font(name="Calibri", bold=True, color=C_HEADER_FG, size=13)
    c.fill = fill(C_HEADER_BG)
    c.alignment = left()
    ws.row_dimensions[1].height = 22

    # ── Row 2: Description ─────────────────────────────────────
    ws.merge_cells("A2:M2")
    c = ws["A2"]
    c.value = island["desc"]
    c.font = Font(name="Calibri", italic=True, color="555555", size=9)
    c.fill = fill("F5EFE6")
    c.alignment = left()

    # ── Row 3: Column headers ──────────────────────────────────
    ws.row_dimensions[3].height = 32
    headers = [""] + SEASONS + ["Annual total"] + SEASONS + ["Annual total"]
    groups  = ([""] +
               ["INPUTS (qty consumed per season)"] * 5 +
               ["OUTPUTS (qty produced per season)"] * 5)
    # group row
    ws.merge_cells("B3:F3")
    g_in = ws["B3"]
    g_in.value = "INPUTS  (qty consumed per season)"
    g_in.font = Font(name="Calibri", bold=True, color=C_HEADER_FG, size=10)
    g_in.fill = fill("7A4A2C")
    g_in.alignment = centre()

    ws.merge_cells("G3:G3")
    ws["G3"].value = "→ Total In"
    ws["G3"].fill = fill("7A4A2C")
    ws["G3"].font = Font(name="Calibri", bold=True, color=C_HEADER_FG, size=9)
    ws["G3"].alignment = centre()

    ws.merge_cells("H3:L3")
    g_out = ws["H3"]
    g_out.value = "OUTPUTS  (qty produced per season)"
    g_out.font = Font(name="Calibri", bold=True, color=C_HEADER_FG, size=10)
    g_out.fill = fill("2E5C38")
    g_out.alignment = centre()

    ws.merge_cells("M3:M3")
    ws["M3"].value = "→ Total Out"
    ws["M3"].fill = fill("2E5C38")
    ws["M3"].font = Font(name="Calibri", bold=True, color=C_HEADER_FG, size=9)
    ws["M3"].alignment = centre()

    # season sub-headers (row 4)
    ws.row_dimensions[4].height = 18
    write_cell(ws, 4, 1, "Resource", C_SECTION_BG, bold=True, fg=C_HEADER_FG, align=left())
    for i, s in enumerate(SEASONS):
        write_cell(ws, 4, 2+i, s, C_SEASON_BG[s], bold=True, fg="1C1410")
    write_cell(ws, 4, 6, "Annual", "D4D0C8", bold=True)
    for i, s in enumerate(SEASONS):
        write_cell(ws, 4, 7+i, s, C_SEASON_BG[s], bold=True, fg="1C1410")
    write_cell(ws, 4, 11, "Annual", "D4D0C8", bold=True)
    # value columns
    write_cell(ws, 4, 12, "In value (Dp)", "F0E6D3", bold=True)
    write_cell(ws, 4, 13, "Out value (Dp)", "E0F0E3", bold=True)

    # ── Resource rows ──────────────────────────────────────────
    all_resources = sorted(set(
        r
        for s_data in island["seasons"].values()
        for section in ("inputs", "outputs")
        for r in s_data[section]
    ))

    data_rows: dict[str, dict] = {}
    for res in all_resources:
        data_rows[res] = {"in": {}, "out": {}}
        for s in SEASONS:
            data_rows[res]["in"][s]  = island["seasons"][s]["inputs"].get(res, 0)
            data_rows[res]["out"][s] = island["seasons"][s]["outputs"].get(res, 0)

    row = 5
    for res in all_resources:
        write_cell(ws, row, 1, res, "F5EFE6", align=left())
        in_cells, out_cells = [], []
        for i, s in enumerate(SEASONS):
            in_val = data_rows[res]["in"][s]
            c_in = ws.cell(row=row, column=2+i, value=in_val if in_val else None)
            c_in.fill = fill(C_INPUT_BG if in_val else "FFFFFF")
            c_in.font = body_font()
            c_in.alignment = centre()
            c_in.border = BORDER
            if in_val:
                in_cells.append(c_in.coordinate)

            out_val = data_rows[res]["out"][s]
            c_out = ws.cell(row=row, column=7+i, value=out_val if out_val else None)
            c_out.fill = fill(C_OUTPUT_BG if out_val else "FFFFFF")
            c_out.font = body_font()
            c_out.alignment = centre()
            c_out.border = BORDER
            if out_val:
                out_cells.append(c_out.coordinate)

        # Annual total (sum of 4 seasons)
        in_refs  = "+".join(f"{get_column_letter(2+i)}{row}" for i in range(4))
        out_refs = "+".join(f"{get_column_letter(7+i)}{row}" for i in range(4))
        c_ann_in = ws.cell(row=row, column=6, value=f"={in_refs}")
        c_ann_in.fill = fill(C_CALC_BG); c_ann_in.font = body_font(); c_ann_in.alignment = centre(); c_ann_in.border = BORDER
        c_ann_out = ws.cell(row=row, column=11, value=f"={out_refs}")
        c_ann_out.fill = fill(C_CALC_BG); c_ann_out.font = body_font(); c_ann_out.alignment = centre(); c_ann_out.border = BORDER

        price = BASE_PRICES.get(res, 0)
        # In value = annual_in * price
        c_vali = ws.cell(row=row, column=12, value=f"=F{row}*{price}")
        c_vali.fill = fill("FFF3E0"); c_vali.font = body_font(color="7A4A2C"); c_vali.alignment = centre(); c_vali.border = BORDER; c_vali.number_format = "#,##0"
        c_valo = ws.cell(row=row, column=13, value=f"=K{row}*{price}")
        c_valo.fill = fill("E8F5E9"); c_valo.font = body_font(color="2E5C38"); c_valo.alignment = centre(); c_valo.border = BORDER; c_valo.number_format = "#,##0"
        row += 1

    # ── Seasonal yield section ─────────────────────────────────
    row += 1
    ws.merge_cells(f"A{row}:M{row}")
    hdr = ws[f"A{row}"]
    hdr.value = "SEASONAL YIELD MULTIPLIER  (applies on top of the base quantities above)"
    hdr.font = Font(name="Calibri", bold=True, color=C_HEADER_FG, size=10)
    hdr.fill = fill(C_SECTION_BG)
    hdr.alignment = left()
    row += 1

    write_cell(ws, row, 1, "Yield multiplier", C_YIELD_BG, bold=False, fg="1C1410", align=left())
    sy = island["seasonal_yield"]
    for i, s in enumerate(SEASONS):
        val = sy[s] if sy else 1.0
        c = ws.cell(row=row, column=2+i, value=val)
        c.fill = fill(C_YIELD_BG)
        c.font = body_font(bold=(val != 1.0))
        c.alignment = centre()
        c.border = BORDER
        c.number_format = "0.00"
    avg_refs = "+".join(f"{get_column_letter(2+i)}{row}" for i in range(4))
    c_avg = ws.cell(row=row, column=6, value=f"=({avg_refs})/4")
    c_avg.fill = fill(C_CALC_BG); c_avg.font = body_font(); c_avg.alignment = centre(); c_avg.border = BORDER; c_avg.number_format = "0.00"
    row += 1

    # ── Labour section ─────────────────────────────────────────
    row += 1
    ws.merge_cells(f"A{row}:M{row}")
    hdr2 = ws[f"A{row}"]
    hdr2.value = "LABOUR REQUIREMENTS  (workers needed per season)"
    hdr2.font = Font(name="Calibri", bold=True, color=C_HEADER_FG, size=10)
    hdr2.fill = fill(C_SECTION_BG)
    hdr2.alignment = left()
    row += 1

    for tier in ("Skilled", "Unskilled"):
        write_cell(ws, row, 1, tier, C_LABOUR_BG, align=left())
        for i, s in enumerate(SEASONS):
            val = island["labour"][s].get(tier, 0)
            c = ws.cell(row=row, column=2+i, value=val if val else None)
            c.fill = fill(C_LABOUR_BG if val else "FFFFFF")
            c.font = body_font()
            c.alignment = centre()
            c.border = BORDER
        row += 1

    # ── Notes row ──────────────────────────────────────────────
    row += 1
    ws.merge_cells(f"A{row}:M{row}")
    note = ws[f"A{row}"]
    note.value = (
        "Yellow = editable inputs/outputs  |  Green = editable outputs  |  "
        "Blue = yield multipliers (avg should stay ≈ 1.0)  |  "
        "Pink = labour  |  Grey = calculated"
    )
    note.font = Font(name="Calibri", italic=True, color="777777", size=8)
    note.alignment = left()


def main():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # remove default sheet

    for island in ISLANDS:
        build_island_sheet(wb, island)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Saved → {OUT}")


if __name__ == "__main__":
    main()
